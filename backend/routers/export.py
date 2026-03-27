import json
import io
import urllib.parse
from datetime import datetime
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from database import get_db
from routers.compatibility import check_compatibility

router = APIRouter()

CATEGORY_LABELS = {
    "cpu": "CPU", "gpu": "GPU", "motherboard": "マザーボード",
    "memory": "メモリ", "storage": "ストレージ", "psu": "電源",
    "case": "ケース", "cooler": "CPUクーラー",
}


def _get_build_data(build_id: int) -> dict:
    conn = get_db()
    build = conn.execute("SELECT * FROM builds WHERE id=?", (build_id,)).fetchone()
    if not build:
        conn.close()
        raise HTTPException(404, "構成が見つかりません")
    build = dict(build)

    rows = conn.execute(
        """SELECT bp.*, p.category, p.brand, p.name, p.model, p.specs,
                  p.tdp, p.benchmark_score, p.reference_price, bp.is_used, bp.custom_price
           FROM build_parts bp JOIN parts p ON bp.part_id=p.id
           WHERE bp.build_id=?""",
        (build_id,),
    ).fetchall()

    parts = []
    total = 0
    total_tdp = 0
    by_cat = {}
    for r in rows:
        d = dict(r)
        try:
            d["specs"] = json.loads(d["specs"])
        except Exception:
            d["specs"] = {}
        price = d["custom_price"] if d["custom_price"] else d["reference_price"]
        d["effective_price"] = price
        total += price * d["quantity"]
        total_tdp += (d["tdp"] or 0) * d["quantity"]
        parts.append(d)
        if d["category"] == "storage":
            by_cat.setdefault("storage", []).append(d)
        else:
            by_cat[d["category"]] = d

    compat = check_compatibility(by_cat)
    conn.close()

    return {
        "build": build,
        "parts": parts,
        "total_price": total,
        "total_tdp": total_tdp,
        "compatibility": compat,
    }


# ─────────────────────────────────────────── Excel ───────────────────────────

@router.get("/excel/{build_id}")
def export_excel(build_id: int):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = _get_build_data(build_id)
    build = data["build"]
    parts = data["parts"]
    compat = data["compatibility"]

    wb = Workbook()

    # ── Sheet 1: 構成サマリー ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "構成サマリー"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    section_fill = PatternFill("solid", fgColor="2E86AB")
    section_font = Font(color="FFFFFF", bold=True)
    alt_fill = PatternFill("solid", fgColor="F0F4F8")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def hdr(ws, row, col, val, font=None, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        if font:
            c.font = font
        if fill:
            c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        return c

    def cell(ws, row, col, val, bold=False, fill=None, align="left"):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold)
        if fill:
            c.fill = fill
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = border
        return c

    # Title
    ws1.merge_cells("A1:F1")
    t = ws1["A1"]
    t.value = f"PC構成: {build['name']}"
    t.font = Font(bold=True, size=16, color="1E3A5F")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    ws1["A2"] = f"作成日: {datetime.now().strftime('%Y/%m/%d')}"
    ws1["D2"] = f"用途: {build.get('purpose', '-')}"
    ws1.row_dimensions[2].height = 20

    # Headers row 3
    headers = ["カテゴリ", "ブランド", "モデル", "TDP (W)", "価格 (円)", "状態"]
    for col, h in enumerate(headers, 1):
        hdr(ws1, 3, col, h, header_font, header_fill)
    ws1.row_dimensions[3].height = 22

    row = 4
    for i, p in enumerate(parts):
        f = alt_fill if i % 2 == 0 else None
        cell(ws1, row, 1, CATEGORY_LABELS.get(p["category"], p["category"]), fill=f)
        cell(ws1, row, 2, p["brand"], fill=f)
        cell(ws1, row, 3, f"{p['name']} {p['model']}", fill=f)
        cell(ws1, row, 4, p.get("tdp", 0), fill=f, align="center")
        cell(ws1, row, 5, p["effective_price"], fill=f, align="right")
        cell(ws1, row, 6, "中古" if p.get("is_used") else "新品", fill=f, align="center")
        row += 1

    # Total
    cell(ws1, row, 4, "合計", bold=True, align="right")
    c = ws1.cell(row=row, column=5, value=data["total_price"])
    c.font = Font(bold=True, size=12, color="C0392B")
    c.alignment = Alignment(horizontal="right")
    c.border = border

    row += 2
    cell(ws1, row, 1, "総消費電力 (推定)", bold=True)
    cell(ws1, row, 2, f"{data['total_tdp']} W")
    row += 1
    cell(ws1, row, 1, "推奨電源容量", bold=True)
    cell(ws1, row, 2, f"{int(data['total_tdp'] * 1.3)} W以上")

    # Column widths
    for i, w in enumerate([18, 16, 40, 12, 14, 10], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: 互換性チェック ────────────────────────────────────────
    ws2 = wb.create_sheet("互換性チェック")
    hdr(ws2, 1, 1, "レベル", header_font, header_fill)
    hdr(ws2, 1, 2, "カテゴリ", header_font, header_fill)
    hdr(ws2, 1, 3, "メッセージ", header_font, header_fill)

    level_colors = {"error": "E74C3C", "warning": "F39C12", "ok": "27AE60"}
    for i, issue in enumerate(compat, 2):
        lvl = issue.get("level", "ok")
        fill = PatternFill("solid", fgColor=level_colors.get(lvl, "FFFFFF"))
        c1 = ws2.cell(row=i, column=1, value={"error": "エラー", "warning": "警告", "ok": "OK"}.get(lvl, lvl))
        c1.fill = fill
        c1.font = Font(color="FFFFFF", bold=True)
        c1.alignment = Alignment(horizontal="center")
        c1.border = border
        cell(ws2, i, 2, issue.get("category", ""))
        cell(ws2, i, 3, issue.get("message", ""))

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 70

    # ── Sheet 3: パーツ詳細 ───────────────────────────────────────────
    ws3 = wb.create_sheet("パーツ詳細")
    ws3["A1"] = "パーツ詳細スペック"
    ws3["A1"].font = Font(bold=True, size=13)
    r = 3
    for p in parts:
        ws3.cell(row=r, column=1, value=CATEGORY_LABELS.get(p["category"], p["category"])).font = Font(bold=True)
        ws3.cell(row=r, column=2, value=f"{p['brand']} {p['name']} {p['model']}")
        r += 1
        specs = p.get("specs", {})
        for k, v in specs.items():
            if isinstance(v, list):
                v = ", ".join(str(x) for x in v)
            ws3.cell(row=r, column=2, value=str(k))
            ws3.cell(row=r, column=3, value=str(v))
            r += 1
        r += 1

    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"PC_build_{build['id']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


# ─────────────────────────────────────────── PDF ────────────────────────────

@router.get("/pdf/{build_id}")
def export_pdf(build_id: int):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiMin-W3"))

    data = _get_build_data(build_id)
    build = data["build"]
    parts = data["parts"]
    compat = data["compatibility"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15*mm, rightMargin=15*mm,
                             topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    JP = "HeiseiMin-W3"

    title_style = ParagraphStyle("title", fontName=JP, fontSize=18, spaceAfter=6,
                                  textColor=colors.HexColor("#1E3A5F"))
    h2_style = ParagraphStyle("h2", fontName=JP, fontSize=13, spaceBefore=12, spaceAfter=4,
                               textColor=colors.HexColor("#2E86AB"), fontWeight="bold")
    body_style = ParagraphStyle("body", fontName=JP, fontSize=9, leading=14)

    story = []

    # Title
    story.append(Paragraph(f"PC構成レポート: {build['name']}", title_style))
    story.append(Paragraph(f"作成日: {datetime.now().strftime('%Y年%m月%d日')}　用途: {build.get('purpose', '-')}", body_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1E3A5F")))
    story.append(Spacer(1, 6))

    # Parts table
    story.append(Paragraph("選択パーツ一覧", h2_style))
    tbl_data = [["カテゴリ", "ブランド", "モデル", "TDP(W)", "価格(円)", "状態"]]
    for p in parts:
        tbl_data.append([
            CATEGORY_LABELS.get(p["category"], p["category"]),
            p["brand"],
            f"{p['name']} {p['model']}"[:40],
            str(p.get("tdp", 0)),
            f"¥{p['effective_price']:,}",
            "中古" if p.get("is_used") else "新品",
        ])
    tbl_data.append(["", "", "", "合計", f"¥{data['total_price']:,}", ""])

    col_w = [28*mm, 25*mm, 65*mm, 18*mm, 22*mm, 14*mm]
    tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), JP),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (3, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F0F4F8")]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F4F8")),
        ("FONTSIZE", (3, -1), (4, -1), 10),
        ("TEXTCOLOR", (4, -1), (4, -1), colors.HexColor("#C0392B")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 6))

    # Power summary
    story.append(Paragraph("電力情報", h2_style))
    pwr_data = [
        ["推定総消費電力", f"{data['total_tdp']} W"],
        ["推奨電源容量", f"{int(data['total_tdp'] * 1.3)} W 以上"],
    ]
    pwr_tbl = Table(pwr_data, colWidths=[60*mm, 50*mm])
    pwr_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), JP),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F0F4F8")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(pwr_tbl)
    story.append(Spacer(1, 6))

    # Compatibility
    story.append(Paragraph("互換性チェック結果", h2_style))
    errors = [i for i in compat if i["level"] == "error"]
    warnings = [i for i in compat if i["level"] == "warning"]
    oks = [i for i in compat if i["level"] == "ok"]
    status_text = "✅ 問題なし" if not errors else f"❌ エラー {len(errors)}件"
    story.append(Paragraph(f"ステータス: {status_text}　警告: {len(warnings)}件", body_style))
    story.append(Spacer(1, 4))

    compat_data = [["レベル", "メッセージ"]]
    for issue in compat:
        lvl_label = {"error": "エラー", "warning": "警告", "ok": "OK"}.get(issue["level"], issue["level"])
        compat_data.append([lvl_label, issue["message"]])

    lvl_colors_map = {"エラー": colors.HexColor("#FADBD8"), "警告": colors.HexColor("#FDEBD0"), "OK": colors.HexColor("#D5F5E3")}
    compat_tbl = Table(compat_data, colWidths=[20*mm, 155*mm], repeatRows=1)
    style_cmds = [
        ("FONTNAME", (0, 0), (-1, -1), JP),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E86AB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
    ]
    for i, issue in enumerate(compat, 1):
        lvl = {"error": "エラー", "warning": "警告", "ok": "OK"}.get(issue["level"], "")
        bg = lvl_colors_map.get(lvl, colors.white)
        style_cmds.append(("BACKGROUND", (0, i), (-1, i), bg))
    compat_tbl.setStyle(TableStyle(style_cmds))
    story.append(compat_tbl)

    doc.build(story)
    buf.seek(0)

    filename = f"PC_build_{build['id']}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )
